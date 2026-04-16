# Django - Vistas y utilidades
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator

# Django - Fechas y tiempo
from django.utils import timezone
from django.utils.timezone import localtime

# Django - Base de datos y consultas
from django.contrib import messages
from django.db.models import Sum, F, Q, ExpressionWrapper, DecimalField, Count
from django.db.models.functions import TruncMonth, TruncDay

# Python - Fechas
from datetime import datetime, time, timedelta, timezone as dt_timezone

# Formularios
from .forms import (
    FormularioRegistro,
    ProductoForm,
    CategoriaForm,
    UsuarioCreateForm,
    UsuarioEditForm
)

# Modelos
from .models import (
    Usuario,
    Producto,
    Categoria,
    Inventario,
    Historial,
    NotificacionConfig,
    ConfigSistema,
    NotificacionEnvio
)

# Utilidades
from .utils import enviar_correo, enviar_notificaciones_usuario, obtener_email_admines

# CSV
import csv





from django.http import HttpResponse
from django.contrib.auth.decorators import login_required as require_login
from .models import Producto, Categoria
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font




from functools import wraps
from django.shortcuts import redirect

def require_usuario_logueado(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if 'usuario_id' not in request.session:
            return redirect('login')  # tu URL de login
        return view_func(request, *args, **kwargs)
    return wrapper

@require_usuario_logueado
def exportar_productos(request):
    productos = Producto.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    headers = ['Nombre', 'Código', 'Categoría', 'Cantidad', 'Precio']
    ws.append(headers)
    
    # Encabezados en negrita
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).font = Font(bold=True)
    
    # Datos
    for p in productos:
        ws.append([p.nombre, p.codigo, p.categoria.nombre, p.cantidad, p.precio])
    
    # Ajuste automático de columnas
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    
    wb.save(response)
    return response


@require_usuario_logueado
def exportar_categorias(request):
    categorias = Categoria.objects.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    
    headers = ['Nombre', 'Descripción', 'Estado']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).font = Font(bold=True)
    
    for c in categorias:
        ws.append([c.nombre, c.descripcion, c.estado])
    
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'
    
    wb.save(response)
    return response




def demo_usuario(request):
    # Limpiar sesión real
    request.session.flush()

    # Marcar modo demo
    request.session['usuario_demo'] = True  # ⚡ aquí ponemos la clave correcta
    request.session['usuario_nombre'] = "Visitante Demo"

    return redirect('invenzo:dashboard')


# ================================
# DECORADOR PARA PROTEGER RUTAS SOPORTA DEMO
# ================================
def require_login(view_func):
    def wrapper(request, *args, **kwargs):
        # Permitir si hay usuario real o usuario demo
        if "usuario_id" not in request.session and not request.session.get("usuario_demo", False):
            return redirect('invenzo:login')
        return view_func(request, *args, **kwargs)
    return wrapper


# ================================
# HOME
# ================================
def inicio(request):
    return render(request, 'usuario/home.html')

# ================================
# REGISTRO
# ================================


def registrar_usuario(request):
    if request.method == 'POST':
        form = FormularioRegistro(request.POST)

        if form.is_valid():
            email = form.cleaned_data['email']

            # Verificar si el email ya existe
            if Usuario.objects.filter(email=email).exists():
                form.add_error("email", "Este correo ya está registrado.")
                return render(request, 'usuario/registro.html', {'form': form})

            # Crear el usuario: SIEMPRE administrador
            usuario = Usuario(
                nombre=form.cleaned_data['nombre'],
                email=email,
                rol="administrador",
                estado="activo"
            )

            usuario.password = form.cleaned_data['contraseña']
            usuario.save()

            asunto = "Bienvenido a Invenzo"
            mensaje = f"Hola {usuario.nombre},\n\nTu cuenta ha sido creada correctamente en Invenzo."
            enviar_correo(asunto, mensaje, usuario.email)

            return redirect('invenzo:login')

    else:
        form = FormularioRegistro()

    return render(request, 'usuario/registro.html', {'form': form})




# ================================
# LOGIN
# ================================


def iniciar_sesion(request):
    error = ''

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('contraseña')

        try:
            usuario = Usuario.objects.get(email=email, password=password)

            if usuario.estado != "activo":
                error = "Tu cuenta está desactivada."
                return render(request, 'usuario/login.html', {'error': error})

            request.session['usuario_id'] = usuario.id
            request.session['usuario_nombre'] = usuario.nombre
            request.session['usuario_rol'] = usuario.rol
            request.session['usuario_foto'] = usuario.foto.url if usuario.foto else None

            return redirect('invenzo:dashboard')

        except Usuario.DoesNotExist:
            error = "Correo o contraseña incorrectos"

    return render(request, 'usuario/login.html', {'error': error})

# ================================
# LOGOUT
# ================================
def cerrar_sesion(request):
    request.session.flush()
    return redirect('invenzo:home')

# ================================
# RECUPERAR CONTRASEÑA (BASE)
# ================================
def recuperar_contraseña(request):
    enviado = False
    email = ""

    if request.method == "POST":
        email = request.POST.get("email")
        enviado = True

    return render(request, "usuario/recuperar.html", {
        "enviado": enviado,
        "email": email
    })

# ================================
# DASHBOARD
# ================================




@require_login
def dashboard(request):
    usuario_demo = request.session.get('usuario_demo', False)

    if usuario_demo:
        # =========================
        # DATOS DE DEMO
        # =========================
        productos = [
            {"nombre": "Laptop Eco", "categoria": "Electrónica", "cantidad": 15, "precio": 1500, "fecha_ingreso": datetime.now()},
            {"nombre": "Botella Reutilizable", "categoria": "Hogar", "cantidad": 30, "precio": 20, "fecha_ingreso": datetime.now()},
            {"nombre": "Cuaderno Reciclado", "categoria": "Papelería", "cantidad": 0, "precio": 5, "fecha_ingreso": datetime.now()},
            {"nombre": "Auriculares Bluetooth", "categoria": "Electrónica", "cantidad": 10, "precio": 100, "fecha_ingreso": datetime.now()},
            {"nombre": "Lámpara Solar", "categoria": "Hogar", "cantidad": 8, "precio": 50, "fecha_ingreso": datetime.now()},
        ]
        productos_totales = len(productos)
        productos_bajos = len([p for p in productos if p['cantidad'] <= 5])
        movimientos_hoy = 0
        valor_total = sum(p['precio'] * p['cantidad'] for p in productos)
        productos_recientes = productos
        tendencia_total = 5.0
        tendencia_bajo = -10.0
        tendencia_mov = 0
        tendencia_valor = 8.5
        form = None  # no necesitamos formulario en demo
    else:
        # =========================
        # USUARIO REAL
        # =========================
        usuario = Usuario.objects.get(id=request.session["usuario_id"])

        # =========================
        # ENVIAR NOTIFICACIONES
        # =========================
        asunto = "Notificación de Invenzo"
        mensaje = f"Hola {usuario.nombre}, tienes nuevas notificaciones en Invenzo."
        # enviar_notificaciones_usuario(usuario)

        # =========================
        # FECHAS
        # =========================
        hoy_local = timezone.localdate()
        inicio_local = datetime.combine(hoy_local, time.min)
        fin_local = datetime.combine(hoy_local, time.max)
        inicio_utc = timezone.make_aware(inicio_local).astimezone(dt_timezone.utc)
        fin_utc = timezone.make_aware(fin_local).astimezone(dt_timezone.utc)
        hace_30 = hoy_local - timedelta(days=30)

        # =========================
        # PRODUCTOS
        # =========================
        productos = Producto.objects.filter(estado='activo')
        productos_totales = productos.count()
        productos_bajos = productos.filter(cantidad__lte=F('stock_minimo')).count()

        # =========================
        # MOVIMIENTOS
        # =========================
        movimientos_hoy = Inventario.objects.filter(
            fecha_movimiento__gte=inicio_utc,
            fecha_movimiento__lte=fin_utc,
            producto__estado='activo'
        ).count()
        mov_pasado = Inventario.objects.filter(
            fecha_movimiento__date__lte=hace_30,
            producto__estado='activo'
        ).count()

        # =========================
        # VALOR INVENTARIO
        # =========================
        valor_total = productos.aggregate(
            total=Sum(F('precio') * F('cantidad'))
        )['total'] or 0

        # =========================
        # PRODUCTOS RECIENTES
        # =========================
        productos_recientes = productos.order_by('-fecha_ingreso')[:5]

        # =========================
        # COMPARACIÓN HACE 30 DÍAS
        # =========================
        productos_pasado = Producto.objects.filter(
            estado='activo',
            fecha_ingreso__lte=hace_30
        ).count()
        stock_pasado = Producto.objects.filter(
            estado='activo',
            cantidad__lte=F('stock_minimo'),
            fecha_ingreso__lte=hace_30
        ).count()
        valor_pasado = Producto.objects.filter(
            estado='activo',
            fecha_ingreso__lte=hace_30
        ).aggregate(total=Sum(F('precio') * F('cantidad')))['total'] or 0

        def tendencia(actual, pasado):
            if pasado == 0:
                return 0
            return round(((actual - pasado) / pasado) * 100, 2)

        tendencia_total = tendencia(productos_totales, productos_pasado)
        tendencia_bajo = tendencia(productos_bajos, stock_pasado)
        tendencia_mov = tendencia(movimientos_hoy, mov_pasado)
        tendencia_valor = tendencia(valor_total, valor_pasado)

        form = ProductoForm()

    # =========================
    # RENDER
    # =========================
    return render(request, 'inventario/dashboard.html', {
        'productos': productos,
        'productos_recientes': productos_recientes,
        'total_productos': productos_totales,
        'stock_bajo': productos_bajos,
        'movimientos_hoy': movimientos_hoy,
        'valor_total': valor_total,
        'page_title': 'Dashboard',
        'form': form,
        'tendencia_total': tendencia_total,
        'tendencia_bajo': tendencia_bajo,
        'tendencia_mov': tendencia_mov,
        'tendencia_valor': tendencia_valor,
        'usuario_demo': usuario_demo,  # esto nos sirve en JS
    })





# ================================
# GRÁFICOS (AJUSTADOS PARA SOLO ACTIVOS)
# ================================
def grafico_distribucion_categorias(request):
    categorias = (
        Categoria.objects
        .filter(estado='activo')
        .annotate(total_productos=Count('producto', filter=Q(producto__estado='activo')))
        .order_by('-total_productos')
    )

    labels = [c.nombre for c in categorias]
    valores = [c.total_productos for c in categorias]

    colores = [
        '#3498db', '#2ecc71', '#f1c40f', '#e67e22', '#9b59b6',
        '#1abc9c', '#e74c3c', '#34495e', '#95a5a6', '#7f8c8d'
    ]

    return JsonResponse({
        "labels": labels,
        "datasets": [{
            "label": "Cantidad de productos por categoría",
            "data": valores,
            "backgroundColor": colores[:len(labels)],
            "borderColor": "#2c3e50",
            "borderWidth": 1
        }]
    })

def grafico_movimiento_diario(request):
    entradas = (
        Inventario.objects
        .filter(tipo_movimiento='entrada', producto__estado='activo')
        .annotate(dia=TruncDay('fecha_movimiento'))
        .values('dia')
        .annotate(total=Sum('cantidad'))
        .order_by('dia')
    )

    salidas = (
        Inventario.objects
        .filter(tipo_movimiento='salida', producto__estado='activo')
        .annotate(dia=TruncDay('fecha_movimiento'))
        .values('dia')
        .annotate(total=Sum('cantidad'))
        .order_by('dia')
    )

    dias = sorted({d['dia'] for d in entradas if d['dia']} | {s['dia'] for s in salidas if s['dia']})
    etiquetas = [d.strftime('%d %b') for d in dias]

    def valores_por_dia(lista):
        return [next((x['total'] for x in lista if x['dia'] == d), 0) for d in dias]

    return JsonResponse({
        "labels": etiquetas,
        "datasets": [
            {
                "label": "Entradas",
                "data": valores_por_dia(entradas),
                "borderColor": "#3c69b7",
                "backgroundColor": "rgba(39, 174, 96, 0.2)",
                "fill": True,
                "tension": 0.4,
            },
            {
                "label": "Salidas",
                "data": valores_por_dia(salidas),
                "borderColor": "#736dc1",
                "backgroundColor": "rgba(192, 57, 43, 0.2)",
                "fill": True,
                "tension": 0.4,
            }
        ]
    })

def tendencia_inventario(request):
    movimientos = (
        Inventario.objects
        .filter(producto__estado='activo')
        .annotate(
            mes=TruncMonth('fecha_movimiento'),
            valor_movimiento=ExpressionWrapper(
                F('cantidad') * F('producto__precio'),
                output_field=DecimalField()
            )
        )
        .values('mes')
        .annotate(valor_total=Sum('valor_movimiento'))
        .order_by('mes')
    )

    labels = [
        m['mes'].strftime('%b') if m['mes'] is not None else "N/A"
        for m in movimientos
    ]
    valores = [float(m['valor_total'] or 0) for m in movimientos]

    return JsonResponse({
        "labels": labels,
        "datasets": [{
            "label": "Valor total del inventario por mes",
            "data": valores,
            "borderColor": "#475ba8",
            "backgroundColor": "rgba(39, 174, 96, 0.2)",
            "fill": True,
            "tension": 0.4
        }]
    })

# ================================
# LISTA DE PRODUCTOS
# ================================
@require_login
def productos(request):
    usuario_demo = request.session.get('usuario_demo', False)

    if usuario_demo:
        # =========================
        # Categorías y productos de demo (sin tocar la DB)
        # =========================
        productos_demo = [
            {"nombre": "Laptop Eco", "categoria": {"nombre": "Electrónica"}, "cantidad": 15, "precio": 1500},
            {"nombre": "Botella Reutilizable", "categoria": {"nombre": "Hogar"}, "cantidad": 30, "precio": 20},
            {"nombre": "Cuaderno Reciclado", "categoria": {"nombre": "Papelería"}, "cantidad": 0, "precio": 5},
            {"nombre": "Auriculares Bluetooth", "categoria": {"nombre": "Electrónica"}, "cantidad": 10, "precio": 100},
            {"nombre": "Lámpara Solar", "categoria": {"nombre": "Hogar"}, "cantidad": 8, "precio": 50},
        ]

        context = {
            "productos": productos_demo,
            "page_obj": productos_demo,  # si quieres paginación real, puedes usar Paginator aquí
            "paginator": None,
            "usuario_demo": True,
        }
        return render(request, "inventario/productos.html", context)

    # =========================
    # Usuario real (base de datos)
    # =========================
    productos = Producto.objects.select_related("categoria").filter(estado='activo')
    paginator = Paginator(productos, 10)
    page = request.GET.get("page")
    page_obj = paginator.get_page(page)

    return render(request, "inventario/productos.html", {
        "productos": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "usuario_demo": False,
    })

@require_login
def productos_disponibles(request):
    status = request.GET.get("status", "all")  # 'all', 'low', 'out'
    search = request.GET.get("search", "")

    productos = Producto.objects.filter(estado="activo")
    productos_inactivos = Producto.objects.filter(estado="inactivo")
    categorias = Categoria.objects.filter(estado='activo')

    # Filtrar búsqueda
    if search:
        productos = productos.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search) |
            Q(categoria__nombre__icontains=search)
        )

    # Filtrar por estado de stock
    if status == "low":
        productos = productos.filter(cantidad__lte=F('stock_minimo')).exclude(cantidad=0)
    elif status == "out":
        productos = productos.filter(cantidad=0)

    productos_all_count = Producto.objects.filter(estado="activo").count()
    productos_bajo_count = Producto.objects.filter(estado="activo", cantidad__lte=F('stock_minimo')).exclude(cantidad=0).count()
    productos_sin_count = Producto.objects.filter(estado="activo", cantidad=0).count()


    form = ProductoForm()

    context = {
        'productos': productos,
        'productos_inactivos': productos_inactivos,
        'productos_all_count': productos_all_count,
        'productos_bajo_count': productos_bajo_count,
        'productos_sin_count': productos_sin_count,
        'form': form,
        'categorias': categorias,
    }


    return render(request, 'inventario/gestor_productos.html', context)


@require_login
def productos_inactivos(request):
    productos = Producto.objects.filter(estado="inactivo")
    categorias = Categoria.objects.filter(estado='activo')

    return render(request, 'inventario/productos_inactivos.html', {
        'productos': productos,
        'categorias': categorias,
    })

# activar_producto: redirige a products disponibles para mantener la UI/estilos
@require_login
def activar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.estado = "activo"
    producto.save()

    Historial.objects.create(
        usuario_id=request.session.get("usuario_id"),
        accion=f"Reactivó producto: {producto.nombre} ({producto.codigo})"
    )

    return redirect('invenzo:productos_disponibles')

# eliminar definitivo (borrado permanente)
@require_login
def eliminar_definitivo(request, id):
    producto = get_object_or_404(Producto, id=id)

    nombre = producto.nombre
    codigo = producto.codigo

    # eliminar imagen si existe
    if producto.imagen:
        try:
            producto.imagen.delete(save=False)
        except Exception:
            pass

    producto.delete()

    Historial.objects.create(
        usuario_id=request.session.get("usuario_id"),
        accion=f"Eliminó DEFINITIVAMENTE el producto: {nombre} ({codigo})"
    )

    return redirect('invenzo:productos_disponibles')

# ================================
# AGREGAR PRODUCTO
# ================================
@require_login
def agregar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)

        if form.is_valid():
            producto = form.save(commit=False)

            if request.FILES.get('imagen'):
                producto.imagen = request.FILES['imagen']

            producto.categoria = form.cleaned_data.get('categoria')
            producto.save()

            Inventario.objects.create(
                producto=producto,
                usuario_id=request.session["usuario_id"],
                tipo_movimiento="entrada",
                cantidad=producto.cantidad,
                observacion="Registro inicial"
            )

            Historial.objects.create(
                usuario_id=request.session.get("usuario_id"),
                accion=f"Creó producto: {producto.nombre} ({producto.codigo})"
            )

            return redirect('invenzo:productos_disponibles')

        else:
            print("ERRORES FORM:", form.errors)

    else:
        form = ProductoForm()

    return render(request, 'modals/producto/modal_agregar_producto.html', {'form': form})

# ================================
# EDITAR PRODUCTO
# ================================


@require_login
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == 'POST':
        producto.nombre = request.POST.get('nombre') or producto.nombre
        producto.codigo = request.POST.get('codigo') or producto.codigo
        producto.precio = request.POST.get('precio') or producto.precio
        producto.stock_minimo = request.POST.get('stock_minimo') or producto.stock_minimo
        producto.stock_maximo = request.POST.get('stock_maximo') or producto.stock_maximo
        producto.descripcion = request.POST.get('descripcion') or producto.descripcion

        categoria_id = request.POST.get('categoria')
        if categoria_id:
            producto.categoria = get_object_or_404(Categoria, id=categoria_id)

        # 🔥 1️⃣ PRIORIDAD: eliminar imagen
        if request.POST.get("eliminar_imagen"):
            if producto.imagen:
                producto.imagen.delete(save=False)
                producto.imagen = None

        # 🔥 2️⃣ SOLO si NO marcó eliminar y sube nueva imagen
        elif request.FILES.get('imagen'):
            if producto.imagen:
                producto.imagen.delete(save=False)
            producto.imagen = request.FILES['imagen']

        producto.save()

        Historial.objects.create(
            usuario_id=request.session.get("usuario_id"),
            accion=f"Editó producto: {producto.nombre} ({producto.codigo})"
        )

        return redirect('invenzo:productos_disponibles')

    categorias = Categoria.objects.filter(estado='activo')
    return render(request, 'modals/producto/modal_editar_producto.html', {
        'producto': producto,
        'categorias': categorias,
    })



# ================================
# ELIMINAR PRODUCTO (soft delete)
# ================================
@require_login
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    producto.estado = "inactivo"
    producto.save()

    Historial.objects.create(
        usuario_id=request.session.get("usuario_id"),
        accion=f"Eliminó producto: {producto.nombre} ({producto.codigo})"
    )

    return redirect('invenzo:productos_disponibles')

# ================================
# MOVIMIENTOS DE INVENTARIO
# ================================



@require_login
def registrar_movimiento(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == "POST":
        tipo = request.POST.get("tipo")
        cantidad = int(request.POST.get("cantidad"))
        observacion = request.POST.get("observacion", "")

        if tipo == "entrada":
            producto.cantidad += cantidad
        else:
            # Evitar stock en negativo
            if cantidad > producto.cantidad:
                # Reemplaza messages.error con redirect con query param o flash
                return redirect("invenzo:productos_disponibles")

            producto.cantidad -= cantidad

        producto.save()

        Inventario.objects.create(
            producto=producto,
            usuario_id=request.session["usuario_id"],
            tipo_movimiento=tipo,
            cantidad=cantidad,
            observacion=observacion
        )

        Historial.objects.create(
            usuario_id=request.session.get("usuario_id"),
            accion=f"{tipo.title()} de {cantidad} unidades del producto {producto.nombre} ({producto.codigo})"
        )

        return redirect("invenzo:productos_disponibles")

    return render(request, "inventario/movimiento.html", {
        "producto": producto
    })




# ================================
# CATEGORIAS
# ================================
@require_login
def categorias(request):
    search = request.GET.get("search", "")

    categorias = Categoria.objects.all()

    if search:
        categorias = categorias.filter(nombre__icontains=search)

    return render(request, "inventario/categorias.html", {
        "categorias": categorias,
    })

@require_login
def crear_categoria(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")

        categoria = Categoria.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            estado="activo"
        )

        Historial.objects.create(
            usuario_id=request.session.get("usuario_id"),
            accion=f"Creó categoría: {categoria.nombre}"
        )

    return redirect("invenzo:categorias")

@require_login
def editar_categoria(request, id):
    categoria = get_object_or_404(Categoria, id=id)

    if request.method == 'POST':
        categoria.nombre = request.POST.get("nombre")
        categoria.descripcion = request.POST.get("descripcion")
        categoria.estado = request.POST.get("estado")
        categoria.save()

        Historial.objects.create(
            usuario_id=request.session.get("usuario_id"),
            accion=f"Editó categoría: {categoria.nombre}"
        )

    return redirect("invenzo:categorias")

@require_login
def eliminar_categoria(request, id):
    categoria = get_object_or_404(Categoria, id=id)
    nombre = categoria.nombre
    categoria.delete()

    Historial.objects.create(
        usuario_id=request.session.get("usuario_id"),
        accion=f"Eliminó categoría: {nombre}"
    )

    return redirect("invenzo:categorias")

# ================================
# EXPORTAR PRODUCTOS
# ================================


# ================================
# CONTROL DE INVENTARIO (ajustado: solo activos)
# ================================
@require_login
def control_inventario(request):
    if "usuario_id" not in request.session:
        return redirect("invenzo:login")

    productos = Producto.objects.filter(estado="activo").order_by("nombre")
    movimientos = Inventario.objects.select_related("producto", "usuario").filter(producto__estado='activo').order_by("-fecha_movimiento")[:5]

    mensaje = ""
    error = ""

    if request.method == "POST":
        producto_id = request.POST.get("producto")
        tipo = request.POST.get("tipo_movimiento")
        cantidad = request.POST.get("cantidad")
        observacion = request.POST.get("observacion", "")

        if not producto_id or not cantidad:
            error = "Debes seleccionar un producto y especificar la cantidad."
        else:
            try:
                cantidad = int(cantidad)
                producto = Producto.objects.get(id=producto_id)

                if tipo == "salida" and cantidad > producto.cantidad:
                    error = "No puedes retirar más del stock disponible."
                else:
                    Inventario.objects.create(
                        producto=producto,
                        usuario_id=request.session["usuario_id"],
                        tipo_movimiento=tipo,
                        cantidad=cantidad,
                        observacion=observacion,
                    )

                    Historial.objects.create(
                        usuario_id=request.session.get("usuario_id"),
                        accion=f"{tipo.title()} de {cantidad} unidades del producto {producto.nombre} ({producto.codigo}) desde Control de Inventario"
                    )

                    if tipo == "entrada":
                        producto.cantidad += cantidad
                    else:
                        producto.cantidad -= cantidad
                    producto.save()

                    mensaje = "Movimiento registrado exitosamente."

            except Producto.DoesNotExist:
                error = "El producto seleccionado no existe."
            except ValueError:
                error = "Cantidad inválida."

    contexto = {
        "productos": productos,
        "movimientos": movimientos,
        "mensaje": mensaje,
        "error": error,
    }

    return render(request, "inventario/control_inventario.html", contexto)

# ================================
# HISTORIAL, ALERTAS Y REPONER STOCK
# ================================
@require_login
def historial(request):
    movimientos = Inventario.objects.select_related("producto", "usuario", "producto__categoria").order_by("-fecha_movimiento")

    search = request.GET.get("search", "")
    tipo = request.GET.get("tipo", "")
    categoria = request.GET.get("categoria", "")
    fecha = request.GET.get("fecha", "")

    if search:
        movimientos = movimientos.filter(
            Q(producto__nombre__icontains=search) |
            Q(producto__codigo__icontains=search) |
            Q(observacion__icontains=search)
        )

    if tipo in ["entrada", "salida"]:
        movimientos = movimientos.filter(tipo_movimiento=tipo)

    if categoria:
        movimientos = movimientos.filter(producto__categoria__nombre__icontains=categoria)

    if fecha:
        movimientos = movimientos.filter(fecha_movimiento__date=fecha)

    paginator = Paginator(movimientos, 8)
    page = request.GET.get("page")
    page_obj = paginator.get_page(page)

    context = {
        "movimientos": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "tipos": ["Entrada", "Salida"],
        "categorias": Categoria.objects.all(),
        "query": request.GET,
    }

    return render(request, "inventario/historial.html", context)



@require_login
def alerta_stock(request):
    usuario_demo = request.session.get('usuario_demo', False)

    if usuario_demo:
        # Datos de demo
        productos = [
            {"nombre": "Laptop Eco", "cantidad": 0, "stock_minimo": 5},
            {"nombre": "Botella Reutilizable", "cantidad": 2, "stock_minimo": 5},
            {"nombre": "Cuaderno Reciclado", "cantidad": 0, "stock_minimo": 3},
        ]
        criticos = [p for p in productos if p['cantidad'] == 0]
        bajos = [p for p in productos if 0 < p['cantidad'] <= p['stock_minimo']]

        context = {
            "productos": productos,
            "page_obj": productos,
            "paginator": None,
            "total_criticos": len(criticos),
            "total_bajos": len(bajos),
            "valor_estimado": sum(p['stock_minimo'] - p['cantidad'] for p in productos if p['cantidad'] <= p['stock_minimo']),
            "query": request.GET,
            "mensaje_alertas": None
        }
        return render(request, "inventario/alerta_stock.html", context)

    # Usuario real
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    config = NotificacionConfig.objects.filter(usuario=usuario).first()

    if not config or not config.alertas_stock:
        context = {
            "productos": [],
            "page_obj": None,
            "paginator": None,
            "total_criticos": 0,
            "total_bajos": 0,
            "total_todos": 0,
            "valor_estimado": 0,
            "query": request.GET,
            "mensaje_alertas": "No tienes activadas las alertas de stock."
        }
        return render(request, "inventario/alerta_stock.html", context)

    productos_activos = Producto.objects.select_related("categoria").filter(estado="activo")
    criticos = productos_activos.filter(cantidad=0)
    bajos = productos_activos.filter(cantidad__gt=0, cantidad__lte=F("stock_minimo"))
    total_criticos = criticos.count()
    total_bajos = bajos.count()
    valor_estimado = sum(
        p.precio * (p.stock_minimo - p.cantidad)
        for p in productos_activos if p.cantidad <= p.stock_minimo
    )

    # Filtros de búsqueda y nivel
    search = request.GET.get("search", "")
    nivel = request.GET.get("nivel", "")
    queryset = productos_activos
    if search:
        queryset = queryset.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search) |
            Q(categoria__nombre__icontains=search)
        )
    if nivel == "critico":
        queryset = criticos
    elif nivel == "bajo":
        queryset = bajos

    paginator = Paginator(queryset, 8)
    page = request.GET.get("page")
    page_obj = paginator.get_page(page)

    # ❌ NO ENVIAR CORREO AQUÍ
    context = {
        "productos": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "total_criticos": total_criticos,
        "total_bajos": total_bajos,
        "total_todos": productos_activos.count(),   # ← ÚNICA LÍNEA AGREGADA
        "valor_estimado": valor_estimado,
        "query": request.GET,
        "mensaje_alertas": None
    }

    return render(request, "inventario/alerta_stock.html", context)





@require_login
def reponer_stock(request, id):
    producto = get_object_or_404(Producto, id=id)

    producto.cantidad = producto.stock_maximo
    producto.save()

    # Registrar movimiento e historial
    Inventario.objects.create(
        producto=producto,
        usuario_id=request.session["usuario_id"],
        tipo_movimiento="entrada",
        cantidad=producto.stock_maximo,
        observacion="Reposición automática desde Alertas de Stock"
    )
    Historial.objects.create(
        usuario_id=request.session.get("usuario_id"),
        accion=f"Reposición automática de {producto.stock_maximo} unidades del producto {producto.nombre} ({producto.codigo})"
    )

    # ✅ Solo aquí se llama a verificar y enviar alerta
    verificar_stock_y_enviar_alerta(producto)

    return redirect("invenzo:alerta_stock")



# ================================
# ADMIN / USUARIOS / CONFIG
# ================================
def require_admin(view):
    def wrapper(request, *args, **kwargs):
        if "usuario_id" not in request.session:
            return redirect('invenzo:login')
        if request.session.get('usuario_rol') != 'administrador':
            return redirect('invenzo:dashboard')
        return view(request, *args, **kwargs)
    return wrapper

@require_admin
def usuarios(request):
    q = request.GET.get('search', '')
    rol = request.GET.get('rol', '')
    estado = request.GET.get('estado', '')

    qs = Usuario.objects.all().order_by('-fecha_creacion')

    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(email__icontains=q)
        )

    if rol:
        qs = qs.filter(rol=rol)
    if estado:
        qs = qs.filter(estado=estado)

    paginator = Paginator(qs, 10)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    context = {
        'usuarios': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'query': request.GET,
    }
    return render(request, 'usuario/usuarios.html', context)

@require_admin
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioCreateForm(request.POST, request.FILES)
        if form.is_valid():
            u = form.save(commit=False)
            u.save()
            return redirect('invenzo:usuarios')
    else:
        form = UsuarioCreateForm()
    return render(request, 'usuario/crear_usuario.html', {'form': form})

@require_admin
def editar_usuario(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    # Evitar que un usuario cambie su propio rol
    if request.session.get('usuario_id') == usuario.id and request.method == "POST":
        nuevo_rol = request.POST.get('rol')
        if nuevo_rol != usuario.rol:
            messages.error(request, "No puedes cambiar tu propio rol.")
            return redirect('invenzo:usuarios')

        # Proteger al último administrador
    admins_activos = Usuario.objects.filter(rol='administrador', estado='activo').count()

    # Si el usuario que se edita es administrador
    es_admin = (usuario.rol == 'administrador')

    # Si es el último admin y están intentando cambiar el rol o desactivarlo
    if es_admin and admins_activos == 1:
        nuevo_rol = request.POST.get('rol')
        nuevo_estado = request.POST.get('estado')

        if nuevo_rol != 'administrador' or nuevo_estado != 'activo':
            messages.error(request, "No puedes modificar ni desactivar al último administrador activo.")
            return redirect('invenzo:usuarios')

    if request.method == 'POST':
        form = UsuarioEditForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('invenzo:usuarios')
    else:
        form = UsuarioEditForm(instance=usuario)
    return render(request, 'usuario/editar_usuario.html', {
    'form': form,
    'usuario': usuario,
    'usuarios_admin_activos': admins_activos,  # 🔥 AGREGA ESTO
})


@require_admin
def desactivar_usuario(request, id):
    if request.method != "POST":
        return redirect('invenzo:usuarios')

    usuario = get_object_or_404(Usuario, id=id)

    # Evitar auto-desactivación
    if request.session.get('usuario_id') == usuario.id:
        messages.error(request, "No puedes desactivarte a ti mismo.")
        return redirect('invenzo:usuarios')

    
    admins_activos = Usuario.objects.filter(rol='administrador', estado='activo').count()
    
    if usuario.rol == 'administrador' and admins_activos == 1:
        messages.error(request, "No puedes desactivar al único administrador.")
        return redirect('invenzo:usuarios')

    usuario.estado = 'inactivo'
    usuario.save()
    messages.success(request, "Usuario desactivado correctamente.")
    return redirect('invenzo:usuarios')


@require_admin
def activar_usuario(request, id):
    if request.method != "POST":
        return redirect('invenzo:usuarios')

    usuario = get_object_or_404(Usuario, id=id)

    # Evitar auto-activación (no debería pasar, pero por seguridad)
    if request.session.get('usuario_id') == usuario.id:
        messages.error(request, "No puedes modificar tu propio estado.")
        return redirect('invenzo:usuarios')

    usuario.estado = 'activo'
    usuario.save()
    messages.success(request, "Usuario activado correctamente.")
    return redirect('invenzo:usuarios')


@require_admin
def reset_password(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    if request.method == 'POST':
        nueva = request.POST.get('password')
        usuario.password = nueva
        usuario.save()
        return redirect('invenzo:usuarios')
    return render(request, 'usuario/reset_password.html', {'usuario': usuario})


# ============================
#  CONFIGURACIÓN
# ============================

@require_login
def configuracion(request):
    usuario_demo = request.session.get('usuario_demo', False)

    if usuario_demo:
        # Datos de demo
        context = {
            "usuario": {
                "nombre_usuario": "DemoUser",
                "email": "demo@invenzo.com",
                "rol": "administrador"
            },
            "mensaje": "Estás en modo demo, no se pueden guardar cambios."
        }
        return render(request, "configuracion/configuracion.html", context)

    # ======================
    # Usuario real
    # ======================
    usuario = Usuario.objects.get(id=request.session["usuario_id"])

    if request.method == "POST":
        nombre_usuario = request.POST.get("nombre_usuario")
        email = request.POST.get("email")
        # Aquí puedes agregar más campos de configuración
        usuario.nombre_usuario = nombre_usuario
        usuario.email = email
        usuario.save()
        mensaje = "Configuración actualizada correctamente."
    else:
        mensaje = None

    context = {
        "usuario": usuario,
        "mensaje": mensaje
    }

    return render(request, "configuracion/configuracion.html", context)

def configuracion_perfil(request):
    usuario = Usuario.objects.get(id=request.session["usuario_id"])

    if request.method == "POST":
        accion = request.POST.get("accion")

        # --- Actualizar Foto ---
        if accion == "foto":
            if request.FILES.get("foto"):
                usuario.foto = request.FILES["foto"]
                usuario.save(update_fields=["foto"])
            return redirect("invenzo:configuracion")

        # --- Actualizar Datos Personales ---
        if accion == "perfil":
            usuario.nombre = request.POST.get("nombre")
            usuario.email = request.POST.get("email")

            if request.POST.get("password"):
                usuario.password = request.POST.get("password")

            usuario.save()
            return redirect("invenzo:configuracion")

    return render(request, "configuracion/configuracion_perfil.html", {"usuario": usuario})



def eliminar_foto(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    usuario.foto.delete(save=True)
    return redirect("invenzo:configuracion_perfil")


def configuracion_notificaciones(request):
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    config, creado = NotificacionConfig.objects.get_or_create(usuario=usuario)

    if request.method == "POST":
        config.alertas_stock = "alertas_stock" in request.POST
        config.movimientos = "movimientos" in request.POST
        config.productos_nuevos = "productos_nuevos" in request.POST
        config.correo_alertas = "correo_alertas" in request.POST
        config.correo_movimientos = "correo_movimientos" in request.POST
        config.save()

        return redirect("invenzo:configuracion")

    return render(request, "configuracion/configuracion_notificaciones.html", {"config": config})



@require_login
def obtener_notificaciones(request):
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    config = NotificacionConfig.objects.filter(usuario=usuario).first()

    notificaciones = []

    if config:
        productos = Producto.objects.filter(estado="activo")
        if config.alertas_stock:
            for p in productos:
                if p.cantidad == 0:
                    notificaciones.append(f"Producto crítico: {p.nombre} (0 unidades)")
                elif p.cantidad <= p.stock_minimo:
                    notificaciones.append(f"Producto bajo: {p.nombre} ({p.cantidad} unidades)")

        if config.productos_nuevos:
            recientes = productos.order_by('-fecha_ingreso')[:5]
            for p in recientes:
                notificaciones.append(f"Producto nuevo: {p.nombre}")

        if config.movimientos:
            movimientos = Inventario.objects.filter(
                fecha_movimiento__gte=timezone.now()-timezone.timedelta(days=1)
            ).select_related('producto').order_by('-fecha_movimiento')[:5]
            for m in movimientos:
                notificaciones.append(f"{m.tipo_movimiento.title()} de {m.cantidad} unidades de {m.producto.nombre}")

    # ❌ NO ENVIAR CORREO AQUÍ
    return JsonResponse({"notificaciones": notificaciones})


# ============================
#  CONFIGURACIÓN GLOBAL (SOLO ADMIN)
# ============================
def configuracion_sistema(request):
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    if usuario.rol != "administrador":
        return redirect("invenzo:configuracion")  # 🔒 bloque para auxiliares

    config, creado = ConfigSistema.objects.get_or_create(id=1)

    if request.method == "POST":
        config.stock_min_global = request.POST.get("stock_min_global")
        config.stock_max_global = request.POST.get("stock_max_global")
        config.nombre_sistema = request.POST.get("nombre_sistema")
        config.save()
        return redirect("invenzo:configuracion")

    return render(request, "configuracion/configuracion_sistema.html", {"config": config})


# ============================
#  GESTIÓN DE USUARIOS (ADMIN)
# ============================
def gestion_usuarios(request):
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    if usuario.rol != "administrador":
        return redirect("invenzo:configuracion")  # 🔒 auxiliares no acceden

    usuarios = Usuario.objects.all()
    return render(request, "configuracion/gestion_usuarios.html", {"usuarios": usuarios})


# ============================
#  CONFIGURACIÓN DE ALERTAS DE STOCK
# ============================
def configuracion_alertas(request):
    usuario = Usuario.objects.get(id=request.session["usuario_id"])
    if usuario.rol != "administrador":
        return redirect("invenzo:configuracion") # 🔒 restringido

    config, creado = ConfigSistema.objects.get_or_create(id=1)

    if request.method == "POST":
        config.stock_min_global = request.POST.get("stock_min_global")
        config.stock_max_global = request.POST.get("stock_max_global")
        config.save()
        return redirect("invenzo:configuracion")

    return render(request, "configuracion/configuracion_alertas.html", {"config": config})



def enviar_notificaciones_usuario(usuario):
    config = NotificacionConfig.objects.filter(usuario=usuario).first()
    if not config:
        return

    mensajes = []

    # Alertas de stock
    if config.alertas_stock:
        productos = Producto.objects.filter(estado="activo")
        criticos = productos.filter(cantidad=0)
        bajos = productos.filter(cantidad__gt=0, cantidad__lte=F("stock_minimo"))

        for p in criticos:
            mensajes.append(f"Producto crítico: {p.nombre} ({p.cantidad} unidades)")
        for p in bajos:
            mensajes.append(f"Producto bajo: {p.nombre} ({p.cantidad} unidades)")

    # Productos nuevos
    if config.productos_nuevos:
        recientes = Producto.objects.filter(estado="activo").order_by('-fecha_ingreso')[:5]
        for p in recientes:
            mensajes.append(f"Producto nuevo: {p.nombre}")

    # Movimientos
    if config.movimientos:
        movimientos = Inventario.objects.filter(
            fecha_movimiento__gte=timezone.now()-timezone.timedelta(days=1)
        ).select_related('producto').order_by('-fecha_movimiento')[:5]

        for m in movimientos:
            mensajes.append(f"{m.tipo_movimiento.title()} de {m.cantidad} unidades de {m.producto.nombre}")

    # Enviar correo si hay mensajes
    if mensajes and config.correo_alertas:  # si el usuario activó recibir por correo
        asunto = "Notificaciones de Invenzo"
        cuerpo = "\n".join(mensajes)
        enviar_correo(asunto, cuerpo, usuario.email)


def verificar_stock_y_enviar_alerta(producto):
    """
    Envía un correo solo si el producto está bajo o sin stock
    y aún no se ha notificado. Solo se envía 1 vez hasta que
    se reponga o supere el stock mínimo.
    """
    if producto.estado != 'activo':
        return

    if producto.notificado_bajo_stock:
        return

    destinatarios = obtener_email_admines()  # lista de emails

    # Producto crítico
    if producto.cantidad == 0:
        asunto = f"Producto crítico: {producto.nombre}"
        mensaje = f"El producto '{producto.nombre}' está sin stock."
        for email in destinatarios:
            enviar_correo(asunto, mensaje, email)
        producto.notificado_bajo_stock = True
        producto.save()
        return

    # Producto bajo
    if producto.cantidad <= producto.stock_minimo:
        asunto = f"Producto bajo: {producto.nombre}"
        mensaje = f"El producto '{producto.nombre}' tiene stock bajo ({producto.cantidad} unidades)."
        for email in destinatarios:
            enviar_correo(asunto, mensaje, email)
        producto.notificado_bajo_stock = True
        producto.save()
        return

    # Stock normal → resetear bandera
    if producto.cantidad > producto.stock_minimo and producto.notificado_bajo_stock:
        producto.notificado_bajo_stock = False
        producto.save()



@require_login
def salida_inventario(request, id):
    producto = get_object_or_404(Producto, id=id)
    cantidad_mov = int(request.POST.get('cantidad', 0))
    
    if cantidad_mov > producto.cantidad:
        cantidad_mov = producto.cantidad

    producto.cantidad -= cantidad_mov
    producto.save()

    Inventario.objects.create(
        producto=producto,
        usuario_id=request.session["usuario_id"],
        tipo_movimiento="salida",
        cantidad=cantidad_mov,
        observacion="Salida manual"
    )

    # ✅ Verificar stock crítico/bajo
    verificar_stock_y_enviar_alerta(producto)

    # DEBUG opcional
    print(f"[DEBUG] Salida: {cantidad_mov}, Nuevo stock: {producto.cantidad}")

    return redirect("invenzo:control_inventario")




@require_login
def entrada_inventario(request, id):
    producto = get_object_or_404(Producto, id=id)
    cantidad_mov = int(request.POST.get('cantidad', 0))
    
    producto.cantidad += cantidad_mov
    producto.save()

    Inventario.objects.create(
        producto=producto,
        usuario_id=request.session["usuario_id"],
        tipo_movimiento="entrada",
        cantidad=cantidad_mov,
        observacion="Entrada manual"
    )

    # ✅ Revisa si se debe resetear la bandera de notificación
    verificar_stock_y_enviar_alerta(producto)

    return redirect("invenzo:control_inventario")

def verificar_stock_y_enviar_alerta(producto):
    """
    Envía un correo si el producto está sin stock o bajo stock.
    Se asegura de enviar solo 1 vez hasta que se reponga.
    """
    if producto.estado != 'activo':
        return

    destinatarios = obtener_email_admines()
    if not destinatarios:
        return

    # Estado actual del producto
    critico = producto.cantidad == 0
    bajo = producto.cantidad <= producto.stock_minimo
    
    # 1. Lógica de ENVÍO (Stock bajo o crítico)
    if critico or bajo:
        # 🔑 VERIFICACIÓN CLAVE: Solo enviar si AÚN NO se ha notificado
        if not producto.notificado_bajo_stock:
            
            asunto = f"{'Producto crítico' if critico else 'Producto bajo'}: {producto.nombre}"
            mensaje = f"El producto '{producto.nombre}' {'está sin stock' if critico else f'tiene stock bajo ({producto.cantidad} unidades)'}."

            for email in destinatarios:
                enviar_correo(asunto, mensaje, email)

            # 🔹 Marcar como enviado para evitar envíos repetidos
            producto.notificado_bajo_stock = True
            # Usamos update_fields para mayor eficiencia, solo actualiza el campo de control
            producto.save(update_fields=['notificado_bajo_stock'])
        
        # Si ya se notificó (bandera True), no hacemos nada en este ciclo
        
    # 2. Lógica de RESETEO (Stock normal)
    else:
        # Stock normal → resetear bandera (solo si estaba True)
        if producto.notificado_bajo_stock:
            producto.notificado_bajo_stock = False
            producto.save(update_fields=['notificado_bajo_stock'])